# teacher_portal/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, timedelta, date
from django.db.models import Q, Count, Avg, Sum, Case, When, Value, IntegerField, Max
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
import json
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os
from django.conf import settings

from api.models import *
from .decorators import teacher_required
from education_department.replacement_utils import (
    get_current_week_dates,
    get_teacher_effective_lessons_for_date,
)
from education_department.models import LessonReplacement

# teacher_portal/views.py
def get_teacher_info(user):
    """Получает информацию о учителе и его группах"""
    try:
        profile = TeacherProfile.objects.get(user=user)
    except TeacherProfile.DoesNotExist:
        profile = None
    
    # Группы, где преподаватель - классный руководитель
    curator_groups = StudentGroup.objects.filter(curator=user)
    
    # Группы, которые преподаватель ведет
    teaching_groups = StudentGroup.objects.filter(
        daily_schedules__lessons__teacher=user
    ).distinct()
    
    # Все группы преподавателя (классный руководитель + преподает)
    # Используем union вместо оператора |
    all_groups_qs = StudentGroup.objects.filter(
        id__in=curator_groups.values_list('id', flat=True)
    ).union(
        StudentGroup.objects.filter(
            id__in=teaching_groups.values_list('id', flat=True)
        )
    )
    
    # Конвертируем union QuerySet в список для удобства
    all_groups = list(all_groups_qs)
    
    # Предметы, которые ведет преподаватель
    subjects = Subject.objects.filter(
        subject_teachers__teacher__user=user
    ).distinct()
    
    return {
        'profile': profile,
        'curator_groups': curator_groups,
        'teaching_groups': teaching_groups,
        'all_groups': all_groups,
        'subjects': subjects,
    }


WEEK_DAY_TO_INDEX = {
    'MON': 0,
    'TUE': 1,
    'WED': 2,
    'THU': 3,
    'FRI': 4,
    'SAT': 5,
    'SUN': 6,
}


def _count_weekday_occurrences(start_date, end_date, weekday_index):
    """Count how many times weekday_index appears in inclusive date range."""
    if start_date > end_date:
        return 0

    days_until_weekday = (weekday_index - start_date.weekday()) % 7
    first_match = start_date + timedelta(days=days_until_weekday)
    if first_match > end_date:
        return 0

    return ((end_date - first_match).days // 7) + 1


def _get_teacher_lesson_occurrences_for_group(teacher, group, start_date, end_date):
    """How many teacher lessons this group had in inclusive date range."""
    if not group or start_date > end_date:
        return 0

    lessons = ScheduleLesson.objects.filter(
        teacher=teacher,
        daily_schedule__student_group=group,
        daily_schedule__is_active=True,
        daily_schedule__is_weekend=False,
    ).select_related('daily_schedule')

    lesson_occurrences = 0
    for lesson in lessons:
        weekday_index = WEEK_DAY_TO_INDEX.get(lesson.daily_schedule.week_day)
        if weekday_index is None:
            continue
        lesson_occurrences += _count_weekday_occurrences(start_date, end_date, weekday_index)

    return lesson_occurrences


def _get_expected_attendance_for_group(teacher, group, start_date, end_date):
    """
    Expected attendance entries for group in period:
    number_of_students * number_of_teacher_lessons_in_range.
    """
    if not group or start_date > end_date:
        return 0

    student_count = StudentProfile.objects.filter(student_group=group).count()
    if student_count == 0:
        return 0

    lesson_occurrences = _get_teacher_lesson_occurrences_for_group(
        teacher,
        group,
        start_date,
        end_date,
    )
    return lesson_occurrences * student_count


def _build_attendance_stats(attendance_qs, expected_total):
    """
    Business rule:
    if there is no attendance mark for lesson/student/date, treat it as present.
    """
    aggregate_data = attendance_qs.aggregate(
        absent=Count(Case(When(status='A', then=1))),
        late=Count(Case(When(status='L', then=1))),
    )

    absent = aggregate_data.get('absent') or 0
    late = aggregate_data.get('late') or 0
    present = max(expected_total - absent - late, 0)

    return {
        'total': expected_total,
        'present': present,
        'absent': absent,
        'late': late,
        'present_percentage': round((present / expected_total) * 100, 1) if expected_total > 0 else 0,
    }


def _get_academic_year_start(current_date):
    """Academic year starts on September 1."""
    if current_date.month >= 9:
        return date(current_date.year, 9, 1)
    return date(current_date.year - 1, 9, 1)



@teacher_required
def dashboard(request):
    """Главная страница преподавателя"""
    teacher_info = get_teacher_info(request.user)
    today = timezone.now().date()
    
    # Статистика
    total_students = StudentProfile.objects.filter(
        student_group__in=teacher_info['all_groups']
    ).count()
    
    # Оценки за сегодня
    today_grades = Grade.objects.filter(
        teacher=request.user,
        date=today
    ).count()
    
    # Посещаемость за сегодня.
    # Если отметки нет, считаем, что студент присутствовал.
    today_attendance_total = 0
    today_attendance_present = 0
    for group in teacher_info['all_groups']:
        expected_total = _get_expected_attendance_for_group(
            request.user,
            group,
            today,
            today,
        )
        attendance_qs = Attendance.objects.filter(
            schedule_lesson__teacher=request.user,
            schedule_lesson__daily_schedule__student_group=group,
            date=today,
        )
        group_stats = _build_attendance_stats(attendance_qs, expected_total)
        today_attendance_total += group_stats['total']
        today_attendance_present += group_stats['present']

    today_attendance = round((today_attendance_present / today_attendance_total) * 100, 1) if today_attendance_total > 0 else 0
    
    # Активные ДЗ
    active_homework = Homework.objects.filter(
        schedule_lesson__teacher=request.user,
        due_date__gte=today
    ).count()
    
    # Ожидающие проверки работы
    recent_submissions = HomeworkSubmission.objects.filter(
        homework__schedule_lesson__teacher=request.user,
        homework__due_date__gte=today - timedelta(days=7)
    )

    # Получаем все домашние задания, по которым есть отправки
    submitted_homework_ids = recent_submissions.values_list('homework_id', flat=True)

    # Получаем работы, по которым еще нет оценок типа 'HW'
    pending_submissions = recent_submissions.filter(
        ~Q(id__in=Grade.objects.filter(
            grade_type='HW',
            subject__in=Subject.objects.filter(
                schedule_lessons__homeworks__id__in=submitted_homework_ids
            ),
            student__in=recent_submissions.values_list('student_id', flat=True)
        ).values_list('student_id', flat=True))
    ).count()
    
    # Расписание на сегодня
    today_schedule = []
    if teacher_info['all_groups']:
        today_schedule = get_teacher_effective_lessons_for_date(
            request.user,
            today,
        )
    
    # Ближайшие события (ДЗ на проверку, оценки к выставлению)
    upcoming_homework = Homework.objects.filter(
        schedule_lesson__teacher=request.user,
        due_date__gte=today
    ).select_related('schedule_lesson__subject', 'student_group').order_by('due_date')[:5]
    
    # Последние объявления
    recent_announcements = Announcement.objects.filter(
        author=request.user
    ).order_by('-created_at')[:5]
    
    context = {
        'teacher_info': teacher_info,
        'today': today,
        'stats': {
            'total_students': total_students,
            'today_grades': today_grades,
            'today_attendance': today_attendance,
            'active_homework': active_homework,
            'pending_submissions': pending_submissions,
        },
        'today_schedule': today_schedule,
        'upcoming_homework': upcoming_homework,
        'recent_announcements': recent_announcements,
    }
    
    return render(request, 'teacher_portal/dashboard.html', context)


@teacher_required
def manage_grades(request):
    """Управление оценками"""
    teacher_info = get_teacher_info(request.user)
    
    # Фильтры
    group_id = request.GET.get('group', '')
    subject_id = request.GET.get('subject', '')
    student_id = request.GET.get('student', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Базовый запрос оценок
    grades_qs = Grade.objects.filter(
        teacher=request.user
    ).select_related('student', 'subject', 'schedule_lesson')
    
    # Применяем фильтры
    if group_id:
        grades_qs = grades_qs.filter(
            schedule_lesson__daily_schedule__student_group_id=group_id
        )
    
    if subject_id:
        grades_qs = grades_qs.filter(subject_id=subject_id)
    
    if student_id:
        grades_qs = grades_qs.filter(student_id=student_id)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            grades_qs = grades_qs.filter(date__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            grades_qs = grades_qs.filter(date__lte=date_to_obj)
        except ValueError:
            pass
    
    # Сортируем по дате
    grades_qs = grades_qs.order_by('-date', '-id')
    
    # Пагинация
    page_number = request.GET.get('page', 1)
    paginator = Paginator(grades_qs, 25)
    page_obj = paginator.get_page(page_number)
    
    # Статистика
    grade_stats = grades_qs.aggregate(
        total=Count('id'),
        average=Avg('value'),
        excellent=Count(Case(When(value__gte=4.5, then=1))),
        good=Count(Case(When(value__gte=3.5, value__lt=4.5, then=1))),
        satisfactory=Count(Case(When(value__gte=2.5, value__lt=3.5, then=1))),
        poor=Count(Case(When(value__lt=2.5, then=1))),
    )
    
    # Данные для фильтров
    groups = teacher_info['all_groups']
    subjects = teacher_info['subjects']
    students = User.objects.filter(
        student_profile__student_group__in=teacher_info['all_groups']
    ).distinct().order_by('last_name', 'first_name')
    
    context = {
        'teacher_info': teacher_info,
        'page_obj': page_obj,
        'grade_stats': grade_stats,
        'groups': groups,
        'subjects': subjects,
        'students': students,
        'filters': {
            'group_id': group_id,
            'subject_id': subject_id,
            'student_id': student_id,
            'date_from': date_from,
            'date_to': date_to,
        },
    }
    
    return render(request, 'teacher_portal/grades.html', context)

@teacher_required
def add_grade(request):
    """Добавление новой оценки"""
    teacher_info = get_teacher_info(request.user)
    
    if request.method == 'POST':
        student_id = request.POST.get('student')
        subject_id = request.POST.get('subject')
        value = request.POST.get('value')
        grade_type = request.POST.get('grade_type')
        comment = request.POST.get('comment', '')
        lesson_id = request.POST.get('lesson', '')
        grade_date = request.POST.get('date', timezone.now().date().isoformat())
        
        # Валидация
        errors = []
        
        if not student_id:
            errors.append('Выберите студента')
        if not subject_id:
            errors.append('Выберите предмет')
        if not value:
            errors.append('Введите оценку')
        else:
            try:
                value = float(value)
                if value < 1 or value > 5:
                    errors.append('Оценка должна быть от 1 до 5')
            except ValueError:
                errors.append('Оценка должна быть числом')
        
        if not grade_type:
            errors.append('Выберите тип оценки')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                student = User.objects.get(id=student_id)
                subject = Subject.objects.get(id=subject_id)
                
                # Получаем пара (если указан)
                schedule_lesson = None
                if lesson_id:
                    schedule_lesson = ScheduleLesson.objects.get(id=lesson_id)
                
                # Создаем оценку
                grade = Grade.objects.create(
                    student=student,
                    subject=subject,
                    schedule_lesson=schedule_lesson,
                    teacher=request.user,
                    value=value,
                    grade_type=grade_type,
                    date=grade_date,
                    comment=comment
                )
                
                messages.success(request, f'Оценка {value} успешно выставлена для {student.get_full_name()}')
                return redirect('teacher_portal:grades')  # Исправлено!
                
            except Exception as e:
                messages.error(request, f'Ошибка при сохранении оценки: {str(e)}')
    
    # Данные для формы
    students = User.objects.filter(
        student_profile__student_group__in=teacher_info['all_groups']
    ).distinct().order_by('last_name', 'first_name')
    
    subjects = teacher_info['subjects']
    
    # Исправленная строка - убираем фильтрацию по дате
    recent_lessons = ScheduleLesson.objects.filter(
        teacher=request.user
    ).select_related('subject', 'daily_schedule__student_group').order_by(
        'daily_schedule__week_day', 'lesson_number'
    )[:10]
    
    context = {
        'teacher_info': teacher_info,
        'students': students,
        'subjects': subjects,
        'recent_lessons': recent_lessons,
        'today': timezone.now().date(),
        'grade_types': Grade.GradeType.choices,  # Добавляем типы оценок
    }
    
    return render(request, 'teacher_portal/grade_form.html', context)


@teacher_required
def edit_grade(request, grade_id):
    """Редактирование оценки"""
    grade = get_object_or_404(Grade, id=grade_id, teacher=request.user)
    teacher_info = get_teacher_info(request.user)
    
    if request.method == 'POST':
        value = request.POST.get('value')
        grade_type = request.POST.get('grade_type')
        comment = request.POST.get('comment', '')
        
        # Валидация
        errors = []
        
        if not value:
            errors.append('Введите оценку')
        else:
            try:
                value = float(value)
                if value < 1 or value > 5:
                    errors.append('Оценка должна быть от 1 до 5')
            except ValueError:
                errors.append('Оценка должна быть числом')
        
        if not grade_type:
            errors.append('Выберите тип оценки')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                grade.value = value
                grade.grade_type = grade_type
                grade.comment = comment
                grade.save()
                
                messages.success(request, 'Оценка успешно обновлена')
                return redirect('teacher_portal:grades')  # Исправлено!
                
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении оценки: {str(e)}')
    
    context = {
        'teacher_info': teacher_info,
        'grade': grade,
        'grade_types': Grade.GradeType.choices,  # Добавляем типы оценок
    }
    
    return render(request, 'teacher_portal/grade_form.html', context)


@teacher_required
def delete_homework(request, homework_id):
    """Удаление домашнего задания"""
    homework = get_object_or_404(Homework, id=homework_id, schedule_lesson__teacher=request.user)
    
    if request.method == 'POST':
        title = homework.title
        homework.delete()
        messages.success(request, f'Задание "{title}" удалено')
        return redirect('teacher_portal:homework')
    
    # Если GET запрос - редирект на список
    return redirect('teacher_portal:homework')


@teacher_required
def delete_grade(request, grade_id):
    """Удаление оценки"""
    grade = get_object_or_404(Grade, id=grade_id, teacher=request.user)
    
    if request.method == 'POST':
        student_name = grade.student.get_full_name()
        grade.delete()
        messages.success(request, f'Оценка для {student_name} удалена')
    
    return redirect('teacher_portal:grades')


@teacher_required
def manage_attendance(request):
    """Управление посещаемостью"""
    teacher_info = get_teacher_info(request.user)

    # Фильтры
    group_id = request.GET.get('group', '')
    subject_id = request.GET.get('subject', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # По умолчанию - сегодня
    selected_date = request.GET.get('date', timezone.now().date().isoformat())

    try:
        selected_date_obj = datetime.strptime(selected_date, '%Y-%m-%d').date()
    except ValueError:
        selected_date_obj = timezone.now().date()

    # Получаем "эффективные" пары на выбранную дату с учетом замен
    lessons = get_teacher_effective_lessons_for_date(
        request.user,
        selected_date_obj,
        group_id=group_id or None,
        subject_id=subject_id or None,
    )

    # Получаем студентов для каждого пары
    attendance_data = []
    for lesson in lessons:
        # Получаем студентов группы
        students = list(StudentProfile.objects.filter(
            student_group=lesson.daily_schedule.student_group
        ).select_related('user').order_by('user__last_name', 'user__first_name'))

        # Получаем посещаемость для этого пары
        attendance_records = {
            record.student_id: record
            for record in Attendance.objects.filter(
                schedule_lesson=lesson,
                date=selected_date_obj
            )
        }

        # Если отметки нет, считаем, что студент присутствовал (P).
        # Сохраняем это в БД, чтобы все отчеты и статистика считались одинаково.
        missing_attendance = [
            Attendance(
                student=student_profile.user,
                schedule_lesson=lesson,
                date=selected_date_obj,
                status='P',
            )
            for student_profile in students
            if student_profile.user_id not in attendance_records
        ]
        if missing_attendance:
            Attendance.objects.bulk_create(missing_attendance, ignore_conflicts=True)
            attendance_records = {
                record.student_id: record
                for record in Attendance.objects.filter(
                    schedule_lesson=lesson,
                    date=selected_date_obj
                )
            }

        lesson_data = {
            'lesson': lesson,
            'students': [],
        }

        for student_profile in students:
            student = student_profile.user
            attendance = attendance_records.get(student.id)

            lesson_data['students'].append({
                'student': student,
                'profile': student_profile,
                'attendance': attendance,
                'status': attendance.status if attendance else 'P',
            })

        attendance_data.append(lesson_data)

    # Данные для фильтров
    groups = teacher_info['all_groups']
    subjects = teacher_info['subjects']

    context = {
        'teacher_info': teacher_info,
        'attendance_data': attendance_data,
        'selected_date': selected_date_obj,
        'groups': groups,
        'subjects': subjects,
        'filters': {
            'group_id': group_id,
            'subject_id': subject_id,
            'date_from': date_from,
            'date_to': date_to,
        },
    }

    return render(request, 'teacher_portal/attendance.html', context)

@require_http_methods(["POST"])
@teacher_required
def save_attendance(request):
    """Сохранение посещаемости"""
    print(f"DEBUG: save_attendance called by {request.user.username}")
    
    try:
        # Читаем тело запроса
        body = request.body.decode('utf-8')
        print(f"DEBUG: Request body: {body}")
        
        data = json.loads(body)
        print(f"DEBUG: Parsed data: {data}")
        
        date_str = data.get('date')
        attendance_data = data.get('attendance', {})
        
        print(f"DEBUG: Date: {date_str}")
        print(f"DEBUG: Attendance data: {attendance_data}")
        
        if not date_str:
            print("DEBUG: No date provided")
            return JsonResponse({
                'success': False,
                'error': 'Дата не указана'
            }, status=400)
        
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        saved_count = 0
        lesson_ids = [int(lesson_id) for lesson_id in attendance_data.keys() if str(lesson_id).isdigit()]
        lessons_map = {
            lesson.id: lesson
            for lesson in ScheduleLesson.objects.filter(id__in=lesson_ids).select_related(
                'teacher',
                'daily_schedule__student_group',
            )
        }
        replacements_map = {
            replacement.original_lesson_id: replacement
            for replacement in LessonReplacement.objects.filter(
                replacement_date=date_obj,
                original_lesson_id__in=lesson_ids,
            ).select_related('replacement_teacher')
        }
        
        # Обрабатываем каждый пара
        for lesson_id, student_statuses in attendance_data.items():
            print(f"DEBUG: Processing lesson {lesson_id}")
            try:
                lesson_int_id = int(lesson_id)
                lesson = lessons_map.get(lesson_int_id)
                if lesson is None:
                    print(f"DEBUG: Lesson {lesson_id} does not exist")
                    continue

                replacement = replacements_map.get(lesson.id)
                effective_teacher_id = (
                    replacement.replacement_teacher_id if replacement else lesson.teacher_id
                )
                if effective_teacher_id != request.user.id:
                    print(f"DEBUG: Lesson teacher mismatch")
                    continue
                print(f"DEBUG: Found lesson: {lesson.id} - {lesson.subject.name}")

                # По умолчанию для всех студентов пары выставляем "Присутствовал",
                # если записи еще нет.
                group_student_ids = list(
                    StudentProfile.objects.filter(
                        student_group=lesson.daily_schedule.student_group
                    ).values_list('user_id', flat=True)
                )
                existing_student_ids = set(
                    Attendance.objects.filter(
                        schedule_lesson=lesson,
                        date=date_obj,
                        student_id__in=group_student_ids,
                    ).values_list('student_id', flat=True)
                )
                default_present_records = [
                    Attendance(
                        student_id=student_id,
                        schedule_lesson=lesson,
                        date=date_obj,
                        status='P',
                    )
                    for student_id in group_student_ids
                    if student_id not in existing_student_ids
                ]
                if default_present_records:
                    Attendance.objects.bulk_create(default_present_records, ignore_conflicts=True)
                
                # Обрабатываем каждого студента в парае
                for student_id, status in student_statuses.items():
                    print(f"DEBUG: Processing student {student_id} with status {status}")
                    try:
                        student_id_int = int(student_id)
                        student = User.objects.get(id=student_id_int)
                        print(f"DEBUG: Found student: {student.get_full_name()}")
                        
                        # Проверяем, что студент в группе пары
                        if student.id not in group_student_ids:
                            print(f"DEBUG: Student not in lesson group")
                            continue
                        
                        # Проверяем корректность статуса
                        valid_statuses = ['P', 'A', 'L']
                        if status not in valid_statuses:
                            print(f"DEBUG: Invalid status: {status}")
                            continue
                        
                        print(f"DEBUG: Saving attendance for {student.get_full_name()}")
                        
                        # Обновляем или создаем запись посещаемости
                        attendance, created = Attendance.objects.update_or_create(
                            student=student,
                            schedule_lesson=lesson,
                            date=date_obj,
                            defaults={'status': status}
                        )
                        
                        saved_count += 1
                        print(f"DEBUG: Saved count: {saved_count}")
                        
                    except User.DoesNotExist:
                        print(f"DEBUG: Student {student_id} does not exist")
                        continue
                    except Exception as e:
                        print(f"DEBUG: Error saving attendance for student {student_id}: {str(e)}")
                        continue
                        
            except ValueError:
                print(f"DEBUG: Invalid lesson id: {lesson_id}")
                continue
            except Exception as e:
                print(f"DEBUG: Error processing lesson {lesson_id}: {str(e)}")
                continue
        
        print(f"DEBUG: Total saved: {saved_count}")
        return JsonResponse({
            'success': True,
            'message': f'Посещаемость для {saved_count} студентов сохранена'
        })
        
    except json.JSONDecodeError as e:
        print(f"DEBUG: JSON decode error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Неверный формат данных'
        }, status=400)
    except Exception as e:
        print(f"DEBUG: General exception: {str(e)}")
        import traceback
        print(f"DEBUG: Traceback: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@teacher_required
def edit_homework(request, homework_id):
    """Редактирование домашнего задания"""
    homework = get_object_or_404(Homework, id=homework_id, schedule_lesson__teacher=request.user)
    teacher_info = get_teacher_info(request.user)
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        due_date = request.POST.get('due_date')
        due_time = request.POST.get('due_time', '23:59')
        attachment = request.FILES.get('attachment')
        
        # Валидация
        errors = []
        
        if not title:
            errors.append('Введите название задания')
        if not due_date:
            errors.append('Укажите срок сдачи')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Обновляем данные
                homework.title = title
                homework.description = description
                
                # Обновляем дату сдачи
                due_datetime = datetime.strptime(f'{due_date} {due_time}', '%Y-%m-%d %H:%M')
                homework.due_date = due_datetime
                
                # Обновляем файл, если загружен новый
                if attachment:
                    homework.attachment = attachment
                
                homework.save()
                
                messages.success(request, f'Задание "{title}" обновлено')
                return redirect('teacher_portal:homework')
                
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении задания: {str(e)}')
    
    # Данные для формы
    lessons = ScheduleLesson.objects.filter(
        teacher=request.user
    ).select_related('subject', 'daily_schedule__student_group').order_by(
        'daily_schedule__week_day', 'lesson_number'
    )[:50]
    
    groups = teacher_info['all_groups']
    
    context = {
        'teacher_info': teacher_info,
        'homework': homework,
        'lessons': lessons,
        'groups': groups,
        'today': timezone.now().date(),
        'edit_mode': True,  # Флаг для режима редактирования
    }
    
    return render(request, 'teacher_portal/homework_form.html', context)

@teacher_required
def manage_homework(request):
    """Управление домашними заданиями"""
    teacher_info = get_teacher_info(request.user)
    
    # Фильтры
    group_id = request.GET.get('group', '')
    subject_id = request.GET.get('subject', '')
    status_filter = request.GET.get('status', 'active')
    
    # Базовый запрос ДЗ
    homework_qs = Homework.objects.filter(
        schedule_lesson__teacher=request.user
    ).select_related('schedule_lesson__subject', 'student_group').order_by('-created_at')
    
    # Применяем фильтры
    if group_id:
        homework_qs = homework_qs.filter(student_group_id=group_id)
    
    if subject_id:
        homework_qs = homework_qs.filter(schedule_lesson__subject_id=subject_id)
    
    if status_filter == 'active':
        homework_qs = homework_qs.filter(due_date__gte=timezone.now().date())
    elif status_filter == 'overdue':
        homework_qs = homework_qs.filter(due_date__lt=timezone.now().date())
    elif status_filter == 'completed':
        # ДЗ, по которым все сдали работы
        pass  # Можно добавить логику
    
    # Пагинация
    page_number = request.GET.get('page', 1)
    paginator = Paginator(homework_qs, 20)
    page_obj = paginator.get_page(page_number)
    
    # Считаем статистику по каждому ДЗ
    for homework in page_obj:
        homework.submission_count = HomeworkSubmission.objects.filter(
            homework=homework
        ).count()
        
        # ВРЕМЕННО УБИРАЕМ ЭТУ СТРОКУ - НЕТ СВЯЗИ МЕЖДУ GRADE И HOMEWORKSUBMISSION
        # homework.graded_count = Grade.objects.filter(
        #     homework_submission__homework=homework,
        #     grade_type='HW'
        # ).count()
        
        homework.graded_count = 0  # Показываем 0, пока не настроим связь
        
        homework.total_students = StudentProfile.objects.filter(
            student_group=homework.student_group
        ).count()
    
    # Данные для фильтров
    groups = teacher_info['all_groups']
    subjects = teacher_info['subjects']
    
    context = {
        'teacher_info': teacher_info,
        'page_obj': page_obj,
        'groups': groups,
        'subjects': subjects,
        'filters': {
            'group_id': group_id,
            'subject_id': subject_id,
            'status_filter': status_filter,
        },
    }
    
    return render(request, 'teacher_portal/homework.html', context)

@teacher_required
def create_homework(request, homework_id=None):
    """Создание или редактирование домашнего задания"""
    homework = None
    edit_mode = False
    
    # Если передан homework_id - это режим редактирования
    if homework_id:
        homework = get_object_or_404(Homework, id=homework_id, schedule_lesson__teacher=request.user)
        edit_mode = True
    
    teacher_info = get_teacher_info(request.user)
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        due_date = request.POST.get('due_date')
        due_time = request.POST.get('due_time', '23:59')
        attachment = request.FILES.get('attachment')
        
        # Валидация
        errors = []
        
        if not title:
            errors.append('Введите название задания')
        if not due_date:
            errors.append('Укажите срок сдачи')
        
        if edit_mode:
            # В режиме редактирования не требуем пара и класс
            lesson = homework.schedule_lesson
            group = homework.student_group
        else:
            # В режиме создания требуем пара и класс
            lesson_id = request.POST.get('lesson')
            group_id = request.POST.get('group')
            
            if not lesson_id:
                errors.append('Выберите пара')
            if not group_id:
                errors.append('Выберите класс')
            
            if not errors:
                lesson = ScheduleLesson.objects.get(id=lesson_id, teacher=request.user)
                group = StudentGroup.objects.get(id=group_id)
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Создаем дату с временем
                due_datetime = datetime.strptime(f'{due_date} {due_time}', '%Y-%m-%d %H:%M')
                
                if edit_mode:
                    # Обновляем существующее задание
                    homework.title = title
                    homework.description = description
                    homework.due_date = due_datetime
                    
                    if attachment:
                        homework.attachment = attachment
                    
                    homework.save()
                    messages.success(request, f'Задание "{title}" обновлено')
                else:
                    # Создаем новое задание
                    homework = Homework.objects.create(
                        title=title,
                        description=description,
                        schedule_lesson=lesson,
                        student_group=group,
                        due_date=due_datetime,
                    )
                    
                    if attachment:
                        homework.attachment = attachment
                        homework.save()
                    
                    messages.success(request, f'Задание "{title}" создано')
                
                return redirect('teacher_portal:homework')
                
            except Exception as e:
                messages.error(request, f'Ошибка: {str(e)}')
    
    # Данные для формы
    lessons = ScheduleLesson.objects.filter(
        teacher=request.user
    ).select_related('subject', 'daily_schedule__student_group').order_by(
        'daily_schedule__week_day', 'lesson_number'
    )[:50]
    
    groups = teacher_info['all_groups']
    
    context = {
        'teacher_info': teacher_info,
        'lessons': lessons,
        'groups': groups,
        'homework': homework,
        'today': timezone.now().date(),
        'default_due_date': (timezone.now() + timedelta(days=7)).date(),
    }
    
    return render(request, 'teacher_portal/homework_form.html', context)


@teacher_required
def homework_submissions(request, homework_id):
    """Проверка работ по домашнему заданию"""
    homework = get_object_or_404(Homework, id=homework_id, schedule_lesson__teacher=request.user)
    teacher_info = get_teacher_info(request.user)
    
    # Получаем все отправки
    submissions = HomeworkSubmission.objects.filter(
        homework=homework
    ).select_related('student').order_by('submitted_at')
    
    # Получаем оценки за домашние работы
    submission_grades = {}
    for submission in submissions:
        # Ищем оценку для этого студента по этому предмету с типом HW
        grade = Grade.objects.filter(
            student=submission.student,
            subject=homework.schedule_lesson.subject,
            schedule_lesson=homework.schedule_lesson,
            grade_type='HW',
            date__gte=submission.submitted_at.date()
        ).order_by('-date', '-id').first()
        
        if grade:
            submission_grades[submission.id] = grade
    
    # Получаем всех студентов группы
    all_students = StudentProfile.objects.filter(
        student_group=homework.student_group
    ).select_related('user').order_by('user__last_name', 'user__first_name')
    
    # Создаем полный список
    students_data = []
    for student_profile in all_students:
        student = student_profile.user
        
        # Ищем отправку
        submission = None
        for sub in submissions:
            if sub.student == student:
                submission = sub
                break
        
        # Ищем оценку
        grade = None
        if submission and submission.id in submission_grades:
            grade = submission_grades[submission.id]

        submission_file_size = 0
        if submission and submission.submission_file and submission.submission_file.name:
            try:
                submission_file_size = submission.submission_file.size
            except (OSError, ValueError):
                submission_file_size = 0
        
        students_data.append({
            'student': student,
            'profile': student_profile,
            'submission': submission,
            'grade': grade,
            'submission_file_size': submission_file_size,
        })
    
    # Статистика
    total_students = all_students.count()
    submitted_count = submissions.count()
    graded_count = len([s for s in students_data if s['grade']])
    pending_review = submitted_count - graded_count
    missing_count = total_students - submitted_count
    submitted_percentage = round((submitted_count / total_students * 100) if total_students > 0 else 0, 1)
    
    context = {
        'teacher_info': teacher_info,
        'homework': homework,
        'students_data': students_data,
        'total_students': total_students,
        'submitted_count': submitted_count,
        'graded_count': graded_count,
        'pending_review': pending_review,
        'missing_count': missing_count,
        'submitted_percentage': submitted_percentage,
        'today': timezone.now().date(),
    }
    
    return render(request, 'teacher_portal/homework_submissions.html', context)


@require_http_methods(["POST"])
@teacher_required
def update_grade(request, grade_id):
    """Обновление существующей оценки"""
    grade = get_object_or_404(Grade, id=grade_id, teacher=request.user)
    
    try:
        value = float(request.POST.get('value', 0))
        comment = request.POST.get('comment', '')
        
        if value < 1 or value > 5:
            return JsonResponse({'error': 'Оценка должна быть от 1 до 5'}, status=400)
        
        grade.value = value
        grade.comment = comment
        grade.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Оценка обновлена',
            'grade_id': grade.id,
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


from django.http import JsonResponse
from django.views.decorators.http import require_GET

@teacher_required
@require_GET
def student_submission_detail(request, homework_id, student_id):
    """Детальная информация об отправке студента (для AJAX)"""
    homework = get_object_or_404(Homework, id=homework_id, schedule_lesson__teacher=request.user)
    student = get_object_or_404(User, id=student_id)
    
    # Проверяем, что студент в группе ДЗ
    if not StudentProfile.objects.filter(user=student, student_group=homework.student_group).exists():
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    # Получаем отправку (если есть)
    submission = HomeworkSubmission.objects.filter(
        homework=homework,
        student=student
    ).first()
    
    # Получаем оценку (если есть)
    grade = None
    if submission:
        grade = Grade.objects.filter(
            student=student,
            subject=homework.schedule_lesson.subject,
            schedule_lesson=homework.schedule_lesson,
            grade_type='HW',
            date__gte=submission.submitted_at.date()
        ).order_by('-date', '-id').first()
    
    from django.urls import reverse
    
    has_file = bool(submission and submission.submission_file and submission.submission_file.name)
    file_name = None
    file_size = None
    file_url = None
    if has_file:
        file_name = os.path.basename(submission.submission_file.name)
        file_url = reverse('teacher_portal:view_submission_file', args=[submission.id])
        try:
            file_size = submission.submission_file.size
        except (OSError, ValueError):
            file_size = None

    data = {
        'success': True,
        'student_id': student.id,
        'student_name': student.get_full_name(),
        'student_patronymic': student.student_profile.patronymic if hasattr(student, 'student_profile') else '',
        'has_submission': submission is not None,
        'submission_id': submission.id if submission else None,
        'submission_text': submission.submission_text if submission else None,
        'submission_date': submission.submitted_at.strftime('%d.%m.%Y %H:%M') if submission else None,
        'has_file': has_file,
        'file_name': file_name,
        'file_size': file_size,
        'file_url': file_url,
        'has_grade': grade is not None,
        'grade_id': grade.id if grade else None,
        'grade_value': grade.value if grade else None,
        'grade_comment': grade.comment if grade else None,
    }
    
    return JsonResponse(data)

@teacher_required
def find_grade_id(request):
    """Поиск ID оценки по студенту и отправке"""
    student_id = request.GET.get('student_id')
    submission_id = request.GET.get('submission_id')
    
    if not student_id or not submission_id:
        return JsonResponse({'error': 'Не указаны параметры'}, status=400)
    
    try:
        # Получаем отправку
        submission = get_object_or_404(HomeworkSubmission, id=submission_id)
        
        # Проверяем доступ
        if submission.homework.schedule_lesson.teacher != request.user:
            return JsonResponse({'error': 'Доступ запрещен'}, status=403)
        
        # Ищем оценку за эту домашнюю работу
        grade = Grade.objects.filter(
            student_id=student_id,
            subject=submission.homework.schedule_lesson.subject,
            schedule_lesson=submission.homework.schedule_lesson,
            grade_type='HW',
            date__gte=submission.submitted_at.date()
        ).order_by('-date', '-id').first()
        
        if grade:
            return JsonResponse({
                'success': True,
                'grade_id': grade.id,
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Оценка не найдена',
            })
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@teacher_required
def view_submission_file(request, submission_id):
    """Просмотр файла, отправленного студентом"""
    submission = get_object_or_404(HomeworkSubmission, id=submission_id)
    
    # Проверяем, что преподаватель имеет доступ к этой работе
    if submission.homework.schedule_lesson.teacher != request.user:
        messages.error(request, 'Доступ запрещен')
        return redirect('teacher_portal:homework')
    
    # Проверяем, есть ли файл
    if not submission.submission_file:
        messages.error(request, 'Файл не прикреплен')
        return redirect('teacher_portal:homework_submissions', homework_id=submission.homework.id)
    
    try:
        # Получаем путь к файлу
        file_path = submission.submission_file.path
        
        # Проверяем существование файла
        if not os.path.exists(file_path):
            messages.error(request, 'Файл не найден')
            return redirect('teacher_portal:homework_submissions', homework_id=submission.homework.id)
        
        # Определяем тип файла
        file_extension = os.path.splitext(file_path)[1].lower()
        content_types = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.txt': 'text/plain',
            '.zip': 'application/zip',
            '.rar': 'application/vnd.rar',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
        }
        
        content_type = content_types.get(file_extension, 'application/octet-stream')
        
        # Определяем действие
        action = request.GET.get('action', 'view')
        
        if action == 'download':
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type,
                as_attachment=True,
                filename=os.path.basename(file_path)
            )
        else:
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type
            )
            
            if file_extension in ['.jpg', '.jpeg', '.png', '.gif']:
                response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Ошибка при открытии файла: {str(e)}')
        return redirect('teacher_portal:homework_submissions', homework_id=submission.homework.id)

@require_http_methods(["POST"])
@teacher_required
def grade_submission(request, submission_id):
    """Выставление оценки за домашнюю работу"""
    submission = get_object_or_404(HomeworkSubmission, id=submission_id)
    
    # Проверяем, что ДЗ создано этим преподавателем
    if submission.homework.schedule_lesson.teacher != request.user:
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    try:
        value = float(request.POST.get('value', 0))
        comment = request.POST.get('comment', '')
        
        if value < 1 or value > 5:
            return JsonResponse({'error': 'Оценка должна быть от 1 до 5'}, status=400)
        
        existing_grade = Grade.objects.filter(
            student=submission.student,
            subject=submission.homework.schedule_lesson.subject,
            schedule_lesson=submission.homework.schedule_lesson,
            grade_type='HW',
            date__gte=submission.submitted_at.date()
        ).order_by('-date', '-id').first()

        if existing_grade:
            grade = existing_grade
            grade.value = value
            grade.comment = comment
            grade.schedule_lesson = submission.homework.schedule_lesson
            grade.teacher = request.user
            grade.date = timezone.now().date()
            grade.save(update_fields=['value', 'comment', 'schedule_lesson', 'teacher', 'date'])
            message = 'Оценка изменена'
        else:
            grade = Grade.objects.create(
                student=submission.student,
                subject=submission.homework.schedule_lesson.subject,
                schedule_lesson=submission.homework.schedule_lesson,
                teacher=request.user,
                value=value,
                grade_type='HW',
                date=timezone.now().date(),
                comment=comment
            )
            message = 'Оценка выставлена'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'grade': {
                'value': grade.value,
                'comment': grade.comment,
                'date': grade.date.isoformat(),
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@teacher_required
def view_schedule(request):
    """Просмотр расписания"""
    teacher_info = get_teacher_info(request.user)

    # Фильтр по группе
    group_id = request.GET.get('group', '')

    week_dates = get_current_week_dates()
    day_name_map = dict(DailySchedule.WeekDay.choices)

    schedule_by_day = {}
    effective_lessons_all = []

    for day_code, day_date in week_dates.items():
        day_lessons = get_teacher_effective_lessons_for_date(
            request.user,
            day_date,
            group_id=group_id or None,
        )
        if not day_lessons:
            continue

        day_name = day_name_map.get(day_code, day_code)
        schedule_by_day[day_name] = {
            'day_name': day_name,
            'day_code': day_code,
            'date': day_date,
            'lessons': day_lessons,
        }
        effective_lessons_all.extend(day_lessons)

    total_lessons = len(effective_lessons_all)
    unique_groups = len({lesson.daily_schedule.student_group_id for lesson in effective_lessons_all})
    unique_subjects = len({lesson.effective_subject.id for lesson in effective_lessons_all})

    days_with_lessons = len(schedule_by_day)
    average_lessons_per_day = total_lessons / days_with_lessons if days_with_lessons > 0 else 0

    context = {
        'teacher_info': teacher_info,
        'schedule_by_day': schedule_by_day,
        'groups': teacher_info['all_groups'],
        'selected_group': group_id,
        # Статистические данные
        'total_lessons': total_lessons,
        'unique_groups': unique_groups,
        'unique_subjects': unique_subjects,
        'average_lessons_per_day': average_lessons_per_day,
    }

    return render(request, 'teacher_portal/schedule.html', context)

@teacher_required
def edit_announcement(request, announcement_id):
    """Редактирование существующего объявления"""
    announcement = get_object_or_404(Announcement, id=announcement_id, author=request.user)
    teacher_info = get_teacher_info(request.user)
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        group_id = request.POST.get('group', '')
        is_for_all = request.POST.get('is_for_all') == 'on'
        
        # Валидация
        errors = []
        
        if not title:
            errors.append('Введите заголовок')
        if not content:
            errors.append('Введите содержание')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Обновляем объявление
                announcement.title = title
                announcement.content = content
                announcement.is_for_all = is_for_all
                
                # Обновляем группу
                if not is_for_all and group_id:
                    group = StudentGroup.objects.get(id=group_id)
                    announcement.student_group = group
                else:
                    announcement.student_group = None
                
                announcement.save()
                messages.success(request, 'Объявление обновлено')
                return redirect('teacher_portal:announcements')
                
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении объявления: {str(e)}')
    
    context = {
        'teacher_info': teacher_info,
        'announcement': announcement,
        'groups': teacher_info['all_groups'],
        'edit_mode': True,  # Флаг для шаблона
    }
    
    return render(request, 'teacher_portal/announcement_form.html', context)


@teacher_required
def delete_announcement(request, announcement_id):
    """Удаление объявления"""
    announcement = get_object_or_404(Announcement, id=announcement_id, author=request.user)
    
    if request.method == 'POST':
        title = announcement.title
        announcement.delete()
        messages.success(request, f'Объявление "{title}" удалено')
        return redirect('teacher_portal:announcements')
    
    # Если GET запрос - редирект на список
    return redirect('teacher_portal:announcements')


@teacher_required
def manage_announcements(request):
    """Управление объявлениями"""
    teacher_info = get_teacher_info(request.user)
    
    # Фильтры
    group_id = request.GET.get('group', '')
    status_filter = request.GET.get('status', 'all')
    
    # Базовый запрос объявлений
    announcements_qs = Announcement.objects.filter(
        author=request.user
    ).select_related('student_group').order_by('-created_at')
    
    # Применяем фильтры
    if group_id:
        if group_id == 'all':
            announcements_qs = announcements_qs.filter(is_for_all=True)
        else:
            announcements_qs = announcements_qs.filter(student_group_id=group_id)
    
    if status_filter == 'active':
        # Активные (последние 7 дней)
        week_ago = timezone.now() - timedelta(days=7)
        announcements_qs = announcements_qs.filter(created_at__gte=week_ago)
    elif status_filter == 'expired':
        # Старые
        week_ago = timezone.now() - timedelta(days=7)
        announcements_qs = announcements_qs.filter(created_at__lt=week_ago)
    
    # Пагинация
    page_number = request.GET.get('page', 1)
    paginator = Paginator(announcements_qs, 20)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'teacher_info': teacher_info,
        'page_obj': page_obj,
        'groups': teacher_info['all_groups'],
        'filters': {
            'group_id': group_id,
            'status_filter': status_filter,
        },
    }
    
    return render(request, 'teacher_portal/announcements.html', context)


@teacher_required
def create_announcement(request):
    """Создание объявления"""
    teacher_info = get_teacher_info(request.user)
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        group_id = request.POST.get('group', '')
        is_for_all = request.POST.get('is_for_all') == 'on'
        
        # Валидация
        errors = []
        
        if not title:
            errors.append('Введите заголовок')
        if not content:
            errors.append('Введите содержание')
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                # Создаем объявление
                announcement = Announcement.objects.create(
                    title=title,
                    content=content,
                    author=request.user,
                    is_for_all=is_for_all,
                )
                
                # Если не "для всех", привязываем к группе
                if not is_for_all and group_id:
                    group = StudentGroup.objects.get(id=group_id)
                    announcement.student_group = group
                    announcement.save()
                
                messages.success(request, 'Объявление опубликовано')
                return redirect('teacher_announcements')
                
            except Exception as e:
                messages.error(request, f'Ошибка при создании объявления: {str(e)}')
    
    context = {
        'teacher_info': teacher_info,
        'groups': teacher_info['all_groups'],
    }
    
    return render(request, 'teacher_portal/announcement_form.html', context)


@teacher_required
def view_students(request):
    """Список студентов"""
    teacher_info = get_teacher_info(request.user)
    
    # Фильтры
    group_id = request.GET.get('group', '')
    search_query = request.GET.get('search', '')
    
    # Базовый запрос студентов
    students_qs = StudentProfile.objects.filter(
        student_group__in=teacher_info['all_groups']
    ).select_related('user', 'student_group').order_by('user__last_name', 'user__first_name')
    
    # Применяем фильтры
    if group_id:
        students_qs = students_qs.filter(student_group_id=group_id)
    
    if search_query:
        students_qs = students_qs.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(patronymic__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )
    
    # Пагинация
    page_number = request.GET.get('page', 1)
    paginator = Paginator(students_qs, 30)
    page_obj = paginator.get_page(page_number)
    
    # Для каждого студента получаем статистику
    total_attendance_stats = {'present': 0, 'total': 0}
    total_grade_stats = {'sum': 0, 'count': 0}
    attendance_period_end = timezone.now().date()
    attendance_period_start = _get_academic_year_start(attendance_period_end)
    group_expected_lessons_cache = {}
    
    for student_profile in page_obj:
        student = student_profile.user
        
        # Статистика оценок
        grade_stats = Grade.objects.filter(
            student=student,
            teacher=request.user
        ).aggregate(
            total=Count('id'),
            average=Avg('value'),
            latest=Max('date'),
        )
        
        student_profile.grade_stats = grade_stats
        
        if grade_stats['total'] and grade_stats['average']:
            total_grade_stats['sum'] += grade_stats['average'] * grade_stats['total']
            total_grade_stats['count'] += grade_stats['total']
        
        # Статистика посещаемости:
        # отсутствие отметки по паре в периоде = "присутствовал".
        attendance_qs = Attendance.objects.filter(
            student=student,
            schedule_lesson__teacher=request.user,
            date__range=[attendance_period_start, attendance_period_end],
        )
        group_id = student_profile.student_group_id
        if group_id not in group_expected_lessons_cache:
            group_expected_lessons_cache[group_id] = _get_teacher_lesson_occurrences_for_group(
                request.user,
                student_profile.student_group,
                attendance_period_start,
                attendance_period_end,
            )
        expected_total = group_expected_lessons_cache[group_id]
        attendance_stats = _build_attendance_stats(attendance_qs, expected_total)
        
        student_profile.attendance_stats = attendance_stats
        
        # Рассчитываем процент присутствия
        if attendance_stats['total'] and attendance_stats['total'] > 0:
            attendance_stats['present_percentage'] = round(
                (attendance_stats['present'] / attendance_stats['total']) * 100, 
                1
            )
        else:
            attendance_stats['present_percentage'] = 0
        
        total_attendance_stats['present'] += attendance_stats['present'] or 0
        total_attendance_stats['total'] += attendance_stats['total'] or 0
    
    # Рассчитываем общую статистику
    avg_grades = total_grade_stats['sum'] / total_grade_stats['count'] if total_grade_stats['count'] > 0 else 0
    
    if total_attendance_stats['total'] > 0:
        attendance_rate = round((total_attendance_stats['present'] / total_attendance_stats['total']) * 100, 1)
    else:
        attendance_rate = 0
    
    # Считаем активных студентов (тех, у кого были оценки или посещаемость в последние 30 дней)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    active_students = User.objects.filter(
        id__in=students_qs.values_list('user_id', flat=True)
    ).filter(
        Q(grades__teacher=request.user, grades__date__gte=thirty_days_ago) |
        Q(attendances__schedule_lesson__teacher=request.user, attendances__date__gte=thirty_days_ago)
    ).distinct().count()
    
    context = {
        'teacher_info': teacher_info,
        'page_obj': page_obj,
        'groups': teacher_info['all_groups'],
        'filters': {
            'group_id': group_id,
            'search_query': search_query,
        },
        'avg_grades': avg_grades,
        'attendance_rate': attendance_rate,
        'active_students': active_students,
        'attendance_period_start': attendance_period_start,
        'attendance_period_end': attendance_period_end,
    }
    
    return render(request, 'teacher_portal/students.html', context)


@teacher_required
def student_detail(request, student_id):
    """Детальная информация об студенте"""
    student_profile = get_object_or_404(StudentProfile, user_id=student_id)
    
    # Проверяем, что студент в группе преподавателя
    teacher_info = get_teacher_info(request.user)
    if student_profile.student_group not in teacher_info['all_groups']:
        messages.error(request, 'Доступ запрещен')
        return redirect('teacher_students')
    
    student = student_profile.user
    
    # Оценки от этого преподавателя
    grades = Grade.objects.filter(
        student=student,
        teacher=request.user
    ).select_related('subject').order_by('-date')[:20]
    
    # Статистика по оценкам
    grade_stats = Grade.objects.filter(
        student=student,
        teacher=request.user
    ).aggregate(
        total=Count('id'),
        average=Avg('value'),
        excellent=Count(Case(When(value__gte=4.5, then=1))),
        good=Count(Case(When(value__gte=3.5, value__lt=4.5, then=1))),
        satisfactory=Count(Case(When(value__gte=2.5, value__lt=3.5, then=1))),
        poor=Count(Case(When(value__lt=2.5, then=1))),
    )
    
    # Посещаемость
    attendance = Attendance.objects.filter(
        student=student,
        schedule_lesson__teacher=request.user
    ).select_related('schedule_lesson__subject').order_by('-date')[:20]
    
    attendance_period_end = timezone.now().date()
    attendance_period_start = _get_academic_year_start(attendance_period_end)
    attendance_qs = Attendance.objects.filter(
        student=student,
        schedule_lesson__teacher=request.user,
        date__range=[attendance_period_start, attendance_period_end],
    )
    expected_total = _get_teacher_lesson_occurrences_for_group(
        request.user,
        student_profile.student_group,
        attendance_period_start,
        attendance_period_end,
    )
    attendance_stats = _build_attendance_stats(attendance_qs, expected_total)
    
    # Домашние задания
    homework_submissions = HomeworkSubmission.objects.filter(
        student=student,
        homework__schedule_lesson__teacher=request.user
    ).select_related('homework', 'homework__schedule_lesson__subject').order_by('-submitted_at')[:10]
    
    context = {
        'teacher_info': teacher_info,
        'student_profile': student_profile,
        'student': student,
        'grades': grades,
        'grade_stats': grade_stats,
        'attendance': attendance,
        'attendance_stats': attendance_stats,
        'homework_submissions': homework_submissions,
        'attendance_period_start': attendance_period_start,
        'attendance_period_end': attendance_period_end,
    }
    
    return render(request, 'teacher_portal/student_detail.html', context)



@teacher_required
def view_statistics(request):
    """Статистика для преподавателя"""
    teacher_info = get_teacher_info(request.user)
    
    # Получаем период из GET-параметров (по умолчанию 30 дней)
    days = int(request.GET.get('days', 30))
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)
    
    # Фильтр по группе
    group_id = request.GET.get('group', '')
    group_filter = Q()
    selected_group_name = None
    selected_group = None
    if group_id:
        try:
            selected_group = StudentGroup.objects.get(id=group_id)
            group_filter = Q(schedule_lesson__daily_schedule__student_group=selected_group)
            selected_group_name = selected_group.name
        except StudentGroup.DoesNotExist:
            pass
    
    # Все оценки за период с учетом фильтра группы
    all_grades = Grade.objects.filter(
        teacher=request.user,
        date__range=[start_date, end_date]
    ).filter(group_filter)
    
    # Общая статистика по всем оценкам
    grade_stats = all_grades.aggregate(
        total=Count('id'),
        average=Avg('value'),
        excellent=Count(Case(When(value__gte=4.5, then=1))),
        good=Count(Case(When(value__gte=3.5, value__lt=4.5, then=1))),
        satisfactory=Count(Case(When(value__gte=2.5, value__lt=3.5, then=1))),
        poor=Count(Case(When(value__lt=2.5, then=1))),
    )
    
    # Качество знаний
    if grade_stats['total'] and grade_stats['total'] > 0:
        quality_count = grade_stats['excellent'] + grade_stats['good']
        grade_stats['quality_percentage'] = (quality_count / grade_stats['total']) * 100
    else:
        grade_stats['quality_percentage'] = 0
    
    # Статистика по предметам
    grade_stats_by_subject = []
    for subject in teacher_info['subjects']:
        subject_grades = all_grades.filter(subject=subject)
        stats = subject_grades.aggregate(
            total=Count('id'),
            average=Avg('value'),
            excellent=Count(Case(When(value__gte=4.5, then=1))),
            good=Count(Case(When(value__gte=3.5, value__lt=4.5, then=1))),
            satisfactory=Count(Case(When(value__gte=2.5, value__lt=3.5, then=1))),
            poor=Count(Case(When(value__lt=2.5, then=1))),
        )
        
        if stats['total'] and stats['total'] > 0:
            quality_count = stats['excellent'] + stats['good']
            stats['quality_percentage'] = (quality_count / stats['total']) * 100
        else:
            stats['quality_percentage'] = 0
        
        grade_stats_by_subject.append({
            'subject': subject,
            'stats': stats,
        })
    
    # Статистика по посещаемости
    attendance_stats_by_group = []
    groups_to_show = []
    if group_id and selected_group:
        groups_to_show = [selected_group]
    else:
        groups_to_show = teacher_info['all_groups']
    
    for group in groups_to_show:
        if group:
            group_attendance = Attendance.objects.filter(
                schedule_lesson__teacher=request.user,
                schedule_lesson__daily_schedule__student_group=group,
                date__range=[start_date, end_date]
            )

            expected_total = _get_expected_attendance_for_group(
                request.user,
                group,
                start_date,
                end_date,
            )
            stats = _build_attendance_stats(group_attendance, expected_total)
            
            attendance_stats_by_group.append({
                'group': group,
                'stats': stats,
            })
    
    # Статистика по ДЗ
    homework_stats = Homework.objects.filter(
        schedule_lesson__teacher=request.user,
        created_at__range=[start_date, end_date]
    ).aggregate(
        total=Count('id'),
        with_submissions=Count(Case(When(submissions__isnull=False, then=1), distinct=True)),
    )
    
    # Добавляем отдельный запрос для проверенных работ
    graded_hw_count = Grade.objects.filter(
        teacher=request.user,
        grade_type='HW',
        date__range=[start_date, end_date]
    ).count()
    
    homework_stats['graded'] = graded_hw_count
    
    # Еженедельная активность
    weekly_activity = []
    for i in range(4, -1, -1):  # Последние 5 недель
        week_start = end_date - timedelta(days=(i * 7) + 6)
        week_end = end_date - timedelta(days=i * 7)
        
        grades_count = Grade.objects.filter(
            teacher=request.user,
            date__range=[week_start, week_end]
        ).filter(group_filter).count()
        
        attendance_count = Attendance.objects.filter(
            schedule_lesson__teacher=request.user,
            date__range=[week_start, week_end]
        ).filter(group_filter).count()
        
        homework_count = Homework.objects.filter(
            schedule_lesson__teacher=request.user,
            created_at__range=[week_start, week_end]
        ).filter(group_filter).count()
        
        weekly_activity.append({
            'week': f'{week_start:%d.%m} - {week_end:%d.%m}',
            'grades': grades_count,
            'attendance': attendance_count,
            'homework': homework_count,
        })
    
    # Считаем общее количество оценок
    total_grades = all_grades.count()
    
    # Считаем среднюю посещаемость:
    # отсутствие отметки = "присутствовал".
    total_attendance = {
        'total': sum(group_data['stats']['total'] for group_data in attendance_stats_by_group),
        'present': sum(group_data['stats']['present'] for group_data in attendance_stats_by_group),
        'absent': sum(group_data['stats']['absent'] for group_data in attendance_stats_by_group),
        'late': sum(group_data['stats']['late'] for group_data in attendance_stats_by_group),
    }
    
    if total_attendance['total'] > 0:
        avg_attendance = round((total_attendance['present'] / total_attendance['total']) * 100, 1)
    else:
        avg_attendance = 0
    
    # Добавляем недостающие переменные для шаблона
    context = {
        'teacher_info': teacher_info,
        'period': {
            'start': start_date,
            'end': end_date,
        },
        'grade_stats': grade_stats,  # Общая статистика по оценкам
        'grade_stats_by_subject': grade_stats_by_subject,
        'attendance_stats_by_group': attendance_stats_by_group,
        'homework_stats': homework_stats,
        'weekly_activity': weekly_activity,
        'days': days,
        'total_grades': total_grades,
        'avg_attendance': avg_attendance,
        'total_attendance': total_attendance,  # Добавляем для шаблона
        'groups_count': len(groups_to_show),  # Количество классов
        'subjects_count': teacher_info['subjects'].count(),  # Количество предметов
        'selected_group': group_id,  # Выбранная группа
        'selected_group_name': selected_group_name,  # Название выбранной группы
        'groups': teacher_info['all_groups'],  # Все группы для фильтра
    }
    
    return render(request, 'teacher_portal/statistics.html', context)


# teacher_portal/views.py - добавьте эту функцию в конец файла

import os
from django.http import FileResponse, Http404
from django.utils.encoding import smart_str
from django.conf import settings

@teacher_required
def view_homework_file(request, homework_id):
    """Просмотр прикрепленного файла домашнего задания"""
    homework = get_object_or_404(Homework, id=homework_id, schedule_lesson__teacher=request.user)
    
    # Проверяем, есть ли прикрепленный файл
    if not homework.attachment:
        messages.error(request, 'Файл не прикреплен к этому заданию')
        return redirect('teacher_portal:homework_submissions', homework_id=homework_id)
    
    try:
        # Получаем путь к файлу
        file_path = homework.attachment.path
        
        # Проверяем существование файла
        if not os.path.exists(file_path):
            messages.error(request, 'Файл не найден на сервере')
            return redirect('teacher_portal:homework_submissions', homework_id=homework_id)
        
        # Определяем тип файла для правильного Content-Type
        file_extension = os.path.splitext(file_path)[1].lower()
        content_types = {
            '.pdf': 'application/pdf',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.txt': 'text/plain',
            '.zip': 'application/zip',
            '.rar': 'application/vnd.rar',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
        }
        
        content_type = content_types.get(file_extension, 'application/octet-stream')
        
        # Определяем, просмотр или скачивание
        action = request.GET.get('action', 'view')
        
        if action == 'download':
            # Скачивание файла
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type,
                as_attachment=True,  # Заставляем браузер скачать файл
                filename=os.path.basename(file_path)
            )
        else:
            # Просмотр в браузере
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type
            )
            
            # Для изображений добавляем заголовки для правильного отображения
            if file_extension in ['.jpg', '.jpeg', '.png', '.gif']:
                response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Ошибка при открытии файла: {str(e)}')
        return redirect('teacher_portal:homework_submissions', homework_id=homework_id)
    

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os
from django.conf import settings
from django.utils import timezone



@teacher_required
def export_statistics_pdf(request):
    """Экспорт статистики в PDF"""
    # Получаем параметры фильтрации
    group_id = request.GET.get('group', '')
    days = int(request.GET.get('days', 30))
    
    teacher = request.user
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)
    
    # Получаем информацию о учителе
    teacher_info = get_teacher_info(teacher)
    
    # Фильтр по группе
    group_filter = Q()
    selected_group_name = None
    selected_group = None
    if group_id:
        try:
            selected_group = StudentGroup.objects.get(id=group_id)
            group_filter = Q(schedule_lesson__daily_schedule__student_group=selected_group)
            selected_group_name = selected_group.name
        except StudentGroup.DoesNotExist:
            pass
    
    # Оценки за период
    grades = Grade.objects.filter(
        teacher=teacher,
        date__gte=start_date,
        date__lte=end_date
    ).filter(group_filter)
    
    # Статистика по оценкам
    grade_stats = {
        'total': grades.count(),
        'average': grades.aggregate(avg=Avg('value'))['avg'] or 0,
        'excellent': grades.filter(value__gte=4.5).count(),
        'good': grades.filter(value__gte=3.5, value__lt=4.5).count(),
        'satisfactory': grades.filter(value__gte=2.5, value__lt=3.5).count(),
        'poor': grades.filter(value__lt=2.5).count(),
    }
    
    # Качество знаний (процент 4 и 5)
    if grade_stats['total'] > 0:
        quality_count = grade_stats['excellent'] + grade_stats['good']
        grade_stats['quality_percentage'] = (quality_count / grade_stats['total']) * 100
    else:
        grade_stats['quality_percentage'] = 0
    
    # Статистика по предметам
    grade_stats_by_subject = []
    for subject in teacher_info['subjects']:
        subject_grades = grades.filter(subject=subject)
        total = subject_grades.count()
        if total > 0:
            avg = subject_grades.aggregate(avg=Avg('value'))['avg'] or 0
            excellent = subject_grades.filter(value__gte=4.5).count()
            good = subject_grades.filter(value__gte=3.5, value__lt=4.5).count()
            quality_percentage = ((excellent + good) / total) * 100 if total > 0 else 0
            
            grade_stats_by_subject.append({
                'subject': subject,
                'stats': {
                    'total': total,
                    'average': avg,
                    'quality_percentage': quality_percentage,
                }
            })
    
    # Посещаемость по группам
    attendance_stats_by_group = []
    groups_to_show = [selected_group] if selected_group else teacher_info['all_groups']
    for group in groups_to_show:
        if not group:
            continue
        group_attendance = Attendance.objects.filter(
            schedule_lesson__teacher=teacher,
            schedule_lesson__daily_schedule__student_group=group,
            date__gte=start_date,
            date__lte=end_date
        )

        expected_total = _get_expected_attendance_for_group(
            teacher,
            group,
            start_date,
            end_date,
        )
        stats = _build_attendance_stats(group_attendance, expected_total)
        
        attendance_stats_by_group.append({
            'group': group,
            'stats': stats,
        })
    
    # Статистика по ДЗ
    homework_stats = Homework.objects.filter(
        schedule_lesson__teacher=teacher,
        created_at__range=[start_date, end_date]
    ).aggregate(
        total=Count('id'),
        with_submissions=Count(Case(When(submissions__isnull=False, then=1), distinct=True)),
    )
    
    # Добавляем проверенные работы
    graded_hw_count = Grade.objects.filter(
        teacher=teacher,
        grade_type='HW',
        date__range=[start_date, end_date]
    ).count()
    homework_stats['graded'] = graded_hw_count
    
    # Общие метрики
    total_grades = grade_stats['total']
    
    total_attendance = {
        'total': sum(group_data['stats']['total'] for group_data in attendance_stats_by_group),
        'present': sum(group_data['stats']['present'] for group_data in attendance_stats_by_group),
        'absent': sum(group_data['stats']['absent'] for group_data in attendance_stats_by_group),
        'late': sum(group_data['stats']['late'] for group_data in attendance_stats_by_group),
    }
    
    if total_attendance['total'] > 0:
        avg_attendance = round((total_attendance['present'] / total_attendance['total']) * 100, 1)
    else:
        avg_attendance = 0
    
    total_homeworks = homework_stats['total']
    
    # Создаем контекст для отчета
    context = {
        'selected_group_name': selected_group_name,
        'grade_stats': grade_stats,
        'grade_stats_by_subject': grade_stats_by_subject,
        'attendance_stats_by_group': attendance_stats_by_group,
        'homework_stats': homework_stats,
        'total_grades': total_grades,
        'avg_attendance': avg_attendance,
        'total_homeworks': total_homeworks,
    }
    
    # Создаем PDF
    buffer = BytesIO()
    
    # Регистрируем шрифт для кириллицы
    font_path = os.path.join(settings.BASE_DIR, "static", "fonts", "ARIAL.TTF")
    try:
        pdfmetrics.registerFont(TTFont("Arial", font_path))
        base_font = "Arial"
    except:
        base_font = "Helvetica"
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"Отчет по успеваемости - {teacher.get_full_name()}"
    )
    
    styles = getSampleStyleSheet()
    for st in styles.byName.values():
        st.fontName = base_font
    
    # Создаем кастомные стили
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=base_font,
        fontSize=16,
        alignment=1,  # center
        spaceAfter=20
    )
    
    story = []
    
    # Заголовок отчета
    story.append(Paragraph(f"ОТЧЕТ ПО УСПЕВАЕМОСТИ", title_style))
    story.append(Paragraph(f"Преподаватель: {teacher.get_full_name()}", styles['Normal']))
    story.append(Paragraph(f"Период: {start_date.strftime('%d.%m.%Y')} – {end_date.strftime('%d.%m.%Y')}", styles['Normal']))
    if group_id:
        story.append(Paragraph(f"Класс: {selected_group_name}", styles['Normal']))
    story.append(Paragraph(f"Дата формирования: {timezone.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 12 * mm))
    
    # Общая статистика
    data = [
        ["Показатель", "Значение"],
        ["Всего оценок", str(total_grades)],
        ["Средний балл", f"{grade_stats['average']:.2f}"],
        ["Качество знаний", f"{grade_stats['quality_percentage']:.1f}%"],
        ["Посещаемость", f"{avg_attendance}%"],
        ["Домашних заданий", str(total_homeworks)],
    ]
    
    table = Table(data, colWidths=[100 * mm, 60 * mm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), base_font),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f2f4f7")),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d5dd")),
        ('ALIGN', (1, 1), (1, -1), "RIGHT"),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(Paragraph("1. Общая статистика", styles['Heading2']))
    story.append(table)
    story.append(Spacer(1, 8 * mm))
    
    # Статистика по предметам
    if grade_stats_by_subject:
        story.append(Paragraph("2. Статистика по предметам", styles['Heading2']))
        data = [["Предмет", "Оценок", "Средний", "Качество"]]
        for subject_data in grade_stats_by_subject:
            data.append([
                subject_data['subject'].name,
                str(subject_data['stats']['total']),
                f"{subject_data['stats']['average']:.1f}",
                f"{subject_data['stats']['quality_percentage']:.1f}%"
            ])
        
        table = Table(data, colWidths=[70 * mm, 30 * mm, 30 * mm, 30 * mm])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), base_font),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f2f4f7")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d5dd")),
            ('ALIGN', (1, 1), (-1, -1), "RIGHT"),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(table)
        story.append(Spacer(1, 8 * mm))
    
    # Посещаемость по классам
    if attendance_stats_by_group:
        story.append(Paragraph("3. Посещаемость по классам", styles['Heading2']))
        data = [["Класс", "Пар", "Присут.", "Отсут.", "Опозд.", "%"]]
        for group_data in attendance_stats_by_group:
            data.append([
                group_data['group'].name,
                str(group_data['stats']['total']),
                str(group_data['stats']['present']),
                str(group_data['stats']['absent']),
                str(group_data['stats']['late']),
                f"{group_data['stats']['present_percentage']:.1f}%"
            ])
        
        table = Table(data, colWidths=[40 * mm, 20 * mm, 20 * mm, 20 * mm, 20 * mm, 20 * mm])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), base_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#f2f4f7")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d5dd")),
            ('ALIGN', (1, 1), (-1, -1), "RIGHT"),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        story.append(Spacer(1, 8 * mm))
    
    # Подпись
    story.append(Spacer(1, 10 * mm))
    story.append(Paragraph("Подпись преподавателя: ____________________", styles['Normal']))
    
    doc.build(story)
    
    pdf = buffer.getvalue()
    buffer.close()
    
    filename = f"statistics_{teacher.username}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response

@teacher_required
def export_statistics_excel(request):
    """Экспорт статистики в Excel"""
    # Получаем параметры фильтрации
    group_id = request.GET.get('group', '')
    days = int(request.GET.get('days', 30))
    
    teacher = request.user
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)
    
    # Получаем информацию о учителе
    teacher_info = get_teacher_info(teacher)
    
    # Фильтр по группе
    group_filter = Q()
    selected_group_name = None
    selected_group = None
    if group_id:
        try:
            selected_group = StudentGroup.objects.get(id=group_id)
            group_filter = Q(schedule_lesson__daily_schedule__student_group=selected_group)
            selected_group_name = selected_group.name
        except StudentGroup.DoesNotExist:
            pass
    
    # Оценки за период
    grades = Grade.objects.filter(
        teacher=teacher,
        date__gte=start_date,
        date__lte=end_date
    ).filter(group_filter)
    
    # Статистика по оценкам
    grade_stats = {
        'total': grades.count(),
        'average': grades.aggregate(avg=Avg('value'))['avg'] or 0,
        'excellent': grades.filter(value__gte=4.5).count(),
        'good': grades.filter(value__gte=3.5, value__lt=4.5).count(),
        'satisfactory': grades.filter(value__gte=2.5, value__lt=3.5).count(),
        'poor': grades.filter(value__lt=2.5).count(),
    }
    
    # Качество знаний (процент 4 и 5)
    if grade_stats['total'] > 0:
        quality_count = grade_stats['excellent'] + grade_stats['good']
        grade_stats['quality_percentage'] = (quality_count / grade_stats['total']) * 100
    else:
        grade_stats['quality_percentage'] = 0
    
    # Статистика по предметам
    grade_stats_by_subject = []
    for subject in teacher_info['subjects']:
        subject_grades = grades.filter(subject=subject)
        total = subject_grades.count()
        if total > 0:
            avg = subject_grades.aggregate(avg=Avg('value'))['avg'] or 0
            excellent = subject_grades.filter(value__gte=4.5).count()
            good = subject_grades.filter(value__gte=3.5, value__lt=4.5).count()
            quality_percentage = ((excellent + good) / total) * 100 if total > 0 else 0
            
            grade_stats_by_subject.append({
                'subject': subject,
                'stats': {
                    'total': total,
                    'average': avg,
                    'quality_percentage': quality_percentage,
                }
            })
    
    # Посещаемость по группам
    attendance_stats_by_group = []
    groups_to_show = [selected_group] if selected_group else teacher_info['all_groups']
    
    for group in groups_to_show:
        if group:
            group_attendance = Attendance.objects.filter(
                schedule_lesson__teacher=teacher,
                schedule_lesson__daily_schedule__student_group=group,
                date__gte=start_date,
                date__lte=end_date
            )

            expected_total = _get_expected_attendance_for_group(
                teacher,
                group,
                start_date,
                end_date,
            )
            stats = _build_attendance_stats(group_attendance, expected_total)
            
            attendance_stats_by_group.append({
                'group': group,
                'stats': stats,
            })
    
    # Статистика по ДЗ
    homework_stats = Homework.objects.filter(
        schedule_lesson__teacher=teacher,
        created_at__range=[start_date, end_date]
    ).aggregate(
        total=Count('id'),
        with_submissions=Count(Case(When(submissions__isnull=False, then=1), distinct=True)),
    )
    
    # Добавляем проверенные работы
    graded_hw_count = Grade.objects.filter(
        teacher=teacher,
        grade_type='HW',
        date__range=[start_date, end_date]
    ).count()
    homework_stats['graded'] = graded_hw_count
    
    # Общие метрики
    total_grades = grade_stats['total']
    
    total_attendance = {
        'total': sum(group_data['stats']['total'] for group_data in attendance_stats_by_group),
        'present': sum(group_data['stats']['present'] for group_data in attendance_stats_by_group),
        'absent': sum(group_data['stats']['absent'] for group_data in attendance_stats_by_group),
        'late': sum(group_data['stats']['late'] for group_data in attendance_stats_by_group),
    }
    
    if total_attendance['total'] > 0:
        avg_attendance = round((total_attendance['present'] / total_attendance['total']) * 100, 1)
    else:
        avg_attendance = 0
    
    total_homeworks = homework_stats['total']
    
    # Создаем контекст для отчета
    context = {
        'selected_group_name': selected_group_name,
        'grade_stats': grade_stats,
        'grade_stats_by_subject': grade_stats_by_subject,
        'attendance_stats_by_group': attendance_stats_by_group,
        'homework_stats': homework_stats,
        'total_grades': total_grades,
        'avg_attendance': avg_attendance,
        'total_homeworks': total_homeworks,
        'groups_count': len(groups_to_show),
        'subjects_count': teacher_info['subjects'].count(),
    }
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Статистика"
    
    # Стили
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    header_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD')
    )
    
    row = 1
    
    # Заголовок
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="ОТЧЕТ ПО УСПЕВАЕМОСТИ")
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # Информация об отчете
    ws.cell(row=row, column=1, value="Преподаватель:").font = header_font
    ws.cell(row=row, column=2, value=teacher.get_full_name()).font = normal_font
    row += 1
    ws.cell(row=row, column=1, value="Период:").font = header_font
    ws.cell(row=row, column=2, value=f"{start_date.strftime('%d.%m.%Y')} – {end_date.strftime('%d.%m.%Y')}").font = normal_font
    row += 1
    if group_id:
        ws.cell(row=row, column=1, value="Класс:").font = header_font
        ws.cell(row=row, column=2, value=selected_group_name or '').font = normal_font
        row += 1
    ws.cell(row=row, column=1, value="Дата формирования:").font = header_font
    ws.cell(row=row, column=2, value=timezone.now().strftime('%d.%m.%Y %H:%M')).font = normal_font
    row += 2
    
    # Общая статистика
    ws.cell(row=row, column=1, value="1. Общая статистика").font = header_font
    row += 1
    
    headers = ["Показатель", "Значение"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    row += 1
    
    stats_data = [
        ["Всего оценок", str(total_grades)],
        ["Средний балл", f"{grade_stats['average']:.2f}"],
        ["Качество знаний", f"{grade_stats['quality_percentage']:.1f}%"],
        ["Посещаемость", f"{avg_attendance}%"],
        ["Домашних заданий", str(total_homeworks)],
    ]
    
    for item in stats_data:
        for col, value in enumerate(item, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = border
        row += 1
    row += 1
    
    # Статистика по предметам
    if grade_stats_by_subject:
        ws.cell(row=row, column=1, value="2. Статистика по предметам").font = header_font
        row += 1
        
        headers = ["Предмет", "Оценок", "Средний", "Качество"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        for subject_data in grade_stats_by_subject:
            data = [
                subject_data['subject'].name,
                str(subject_data['stats']['total']),
                f"{subject_data['stats']['average']:.1f}",
                f"{subject_data['stats']['quality_percentage']:.1f}%"
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = border
            row += 1
        row += 1
    
    # Посещаемость по классам
    if attendance_stats_by_group:
        ws.cell(row=row, column=1, value="3. Посещаемость по классам").font = header_font
        row += 1
        
        headers = ["Класс", "Пар", "Присут.", "Отсут.", "Опозд.", "%"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        for group_data in attendance_stats_by_group:
            data = [
                group_data['group'].name,
                str(group_data['stats']['total']),
                str(group_data['stats']['present']),
                str(group_data['stats']['absent']),
                str(group_data['stats']['late']),
                f"{group_data['stats']['present_percentage']:.1f}%"
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = border
            row += 1
        row += 2
    
    # Подпись
    ws.cell(row=row, column=1, value="Подпись преподавателя: ____________________").font = normal_font
    
    # Автоширина колонок
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = adjusted_width
    
    # Сохраняем в response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"statistics_{teacher.username}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

